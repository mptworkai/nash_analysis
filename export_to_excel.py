#!/usr/bin/env python3
"""
Export events.db to an Excel workbook, one row per event.

Columns: date, name, series, program, fiscal_year, artists,
         all revenue fields, all expense fields, ticket counts,
         gross margin, and notes.

Usage:
  python export_to_excel.py                        # writes events_export.xlsx
  python export_to_excel.py --db PATH              # use a different db
  python export_to_excel.py --out FILE.xlsx        # specify output path
  python export_to_excel.py --help

Examples:
  python export_to_excel.py
  python export_to_excel.py --db backup.db --out backup_export.xlsx
"""

import sqlite3
import sys
from pathlib import Path

if "-h" in sys.argv or "--help" in sys.argv:
    print(__doc__)
    sys.exit(0)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── column definitions (must match app.py) ───────────────────────────────────
REVENUE_COLS = [
    ("gross_ticket_sales", "Ticket Sales"),
    ("food_beverage",      "F&B Revenue"),
    ("merchandise_sales",  "Merch"),
    ("cd_sales",           "CDs"),
    ("donations",          "Donations"),
    ("grant_underwriting", "Grants"),
    ("venue_rental",       "Venue Rental"),
    ("other_revenue",      "Other Rev"),
]
EXPENSE_COLS = [
    ("performance_fee",       "Performance Fee"),
    ("tech_support",          "Tech Support"),
    ("security",              "Security"),
    ("musicians_travel",      "Musicians Travel"),
    ("equipment_rental",      "Equipment Rental"),
    ("advertising",           "Advertising"),
    ("misc_expense",          "Misc Expense"),
    ("merchant_fees",         "Merchant Fees"),
    ("food_beverage_expense", "F&B Expense"),
]
COUNT_COLS = [
    ("tickets_sold",      "Tickets Sold"),
    ("full_price_tickets","Full Price"),
    ("discount_tickets",  "Discount"),
    ("comp_tickets",      "Comp"),
]

# ── argument parsing ──────────────────────────────────────────────────────────
db_path  = Path("events.db")
out_path = Path("events_export.xlsx")
args = sys.argv[1:]
i = 0
while i < len(args):
    if args[i] == "--db" and i + 1 < len(args):
        db_path = Path(args[i + 1]); i += 2
    elif args[i] == "--out" and i + 1 < len(args):
        out_path = Path(args[i + 1]); i += 2
    else:
        i += 1

if not db_path.exists():
    print(f"Database not found: {db_path}", file=sys.stderr)
    sys.exit(1)

# ── query ─────────────────────────────────────────────────────────────────────
con = sqlite3.connect(db_path)
con.row_factory = sqlite3.Row

events = con.execute("""
    SELECT e.*,
           GROUP_CONCAT(a.name, ', ') AS artists
    FROM event e
    LEFT JOIN event_artist ea ON ea.event_id = e.event_id
    LEFT JOIN artist a        ON a.artist_id  = ea.artist_id
    GROUP BY e.event_id
    ORDER BY e.date, e.name
""").fetchall()
con.close()

# ── build header + track formula column positions ─────────────────────────────
HEADERS = [
    ("Date",        "date"),
    ("Event",       "name"),
    ("Series",      "series"),
    ("Program",     "program"),
    ("FY",          "fiscal_year"),
    ("Artists",     "artists"),
]
for _, label in REVENUE_COLS:
    HEADERS.append((label, label))
REV_START = 7                            # 1-based col of first revenue field
REV_END   = REV_START + len(REVENUE_COLS) - 1
HEADERS.append(("Total Revenue", "__rev"))
TOT_REV_COL = REV_END + 1

for _, label in EXPENSE_COLS:
    HEADERS.append((label, label))
EXP_START = TOT_REV_COL + 1
EXP_END   = EXP_START + len(EXPENSE_COLS) - 1
HEADERS.append(("Total Expense", "__exp"))
TOT_EXP_COL = EXP_END + 1

HEADERS.append(("Gross Margin",  "__margin"))
MARGIN_COL = TOT_EXP_COL + 1

for _, label in COUNT_COLS:
    HEADERS.append((label, label))
HEADERS.append(("Notes", "notes"))

# ── styles ────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
REV_FILL   = PatternFill("solid", fgColor="D1FAE5")
EXP_FILL   = PatternFill("solid", fgColor="FEE2E2")
TOT_FONT   = Font(bold=True, size=10)
NEG_FONT   = Font(bold=True, color="DC2626", size=10)
MONEY_FMT  = '#,##0.00'
thin       = Side(style="thin", color="D1D5DB")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Events"
ws.freeze_panes = "A2"

# header row
for col_idx, (label, _) in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border    = BORDER

# build label→db-key lookup
rev_labels  = {label for _, label in REVENUE_COLS}
exp_labels  = {label for _, label in EXPENSE_COLS}
count_labels= {label for _, label in COUNT_COLS}
rev_db      = {label: col for col, label in REVENUE_COLS}
exp_db      = {label: col for col, label in EXPENSE_COLS}
count_db    = {label: col for col, label in COUNT_COLS}

rev_start_letter = get_column_letter(REV_START)
rev_end_letter   = get_column_letter(REV_END)
exp_start_letter = get_column_letter(EXP_START)
exp_end_letter   = get_column_letter(EXP_END)
tot_rev_letter   = get_column_letter(TOT_REV_COL)
tot_exp_letter   = get_column_letter(TOT_EXP_COL)

for row_idx, ev in enumerate(events, 2):
    for col_idx, (label, key) in enumerate(HEADERS, 1):
        if key == "__rev":
            val = f"=SUM({rev_start_letter}{row_idx}:{rev_end_letter}{row_idx})"
        elif key == "__exp":
            val = f"=SUM({exp_start_letter}{row_idx}:{exp_end_letter}{row_idx})"
        elif key == "__margin":
            val = f"={tot_rev_letter}{row_idx}-{tot_exp_letter}{row_idx}"
        elif key in rev_labels:
            val = ev[rev_db[key]] or 0
        elif key in exp_labels:
            val = ev[exp_db[key]] or 0
        elif key in count_labels:
            val = ev[count_db[key]] or 0
        else:
            val = ev[key] if key in ev.keys() else None

        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top")

        if key in rev_labels or key == "__rev":
            cell.number_format = MONEY_FMT
            cell.fill = REV_FILL
            if key == "__rev":
                cell.font = TOT_FONT
        elif key in exp_labels or key == "__exp":
            cell.number_format = MONEY_FMT
            cell.fill = EXP_FILL
            if key == "__exp":
                cell.font = TOT_FONT
        elif key == "__margin":
            cell.number_format = MONEY_FMT
            cell.font = TOT_FONT
        elif key in count_labels:
            cell.number_format = '#,##0'

# column widths
COL_WIDTHS = {"Date": 12, "Event": 32, "Series": 16, "Program": 14,
              "FY": 5, "Artists": 28, "Notes": 30,
              "Total Revenue": 14, "Total Expense": 14, "Gross Margin": 13}
for col_idx, (label, _) in enumerate(HEADERS, 1):
    w = COL_WIDTHS.get(label, 13)
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 28

wb.save(out_path)
print(f"Exported {len(events)} events → {out_path}")
