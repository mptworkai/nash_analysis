#!/usr/bin/env python3
"""
Unpivot anonymized Excel files into a SQLite database.

Reads anonymized xlsx files from mod_excel/, detects person+date columns,
and loads tall-format rows into a SQLite database.

Expected column header format: "Person_001 November 6, 2025"
Non-person columns (row labels) are preserved as additional fields.

Usage:
  python load_to_sqlite.py                        # reads mod_excel/*.xlsx
  python load_to_sqlite.py --db mydata.db         # specify db name
  python load_to_sqlite.py --help
"""

import sys
import sqlite3
from difflib import get_close_matches
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Missing dependency. Run: pip install openpyxl")
    sys.exit(1)

MONTH_NAMES = [
    "january","february","march","april","may","june",
    "july","august","september","october","november","december"
]

# canonical spellings for correcting typos before date parsing
MONTH_CANONICAL = {m: m.capitalize() for m in MONTH_NAMES}

def is_month(token: str) -> bool:
    """Match month names including typos like Februrary, Novmeber, Januray."""
    clean = token.lower().rstrip(",").strip("()")
    if clean in MONTH_NAMES:
        return True
    return bool(get_close_matches(clean, MONTH_NAMES, n=1, cutoff=0.8))

def fix_month_typos(date_str: str) -> str:
    """Replace misspelled month names with canonical spellings."""
    tokens = date_str.split()
    fixed = []
    for token in tokens:
        clean = token.lower().rstrip(",").strip("()")
        matches = get_close_matches(clean, MONTH_NAMES, n=1, cutoff=0.8)
        if matches:
            # preserve trailing comma if present
            suffix = "," if token.rstrip().endswith(",") else ""
            fixed.append(matches[0].capitalize() + suffix)
        else:
            fixed.append(token)
    return " ".join(fixed)

HELP = """
Usage:
  python load_to_sqlite.py
  python load_to_sqlite.py --db mydata.db

Description:
  Reads anonymized Excel files from mod_excel/, unpivots person+date
  columns into tall-format rows, and loads into a SQLite database.

  Input:  mod_excel/*_anonymized.xlsx
  Output: mod_excel/financial_data.db  (or --db path)

  Table: daily_values
    source_file  — original filename
    sheet        — tab name (typically the month)
    person       — anonymized person alias (e.g. Person_001)
    date         — ISO date (YYYY-MM-DD)
    value        — numeric cell value
    [label cols] — any non-person columns preserved as extra fields

Options:
  --db <path>   Path for the SQLite database file
                Default: mod_excel/financial_data.db
  --help        Show this help message and exit
"""

def split_name_and_date(header: str):
    tokens = header.split()
    for i, token in enumerate(tokens):
        if is_month(token):
            return " ".join(tokens[:i]).strip(), " ".join(tokens[i:]).strip()
    return header.strip(), ""

def find_header_row(ws) -> int:
    for row in ws.iter_rows():
        if any(cell.value is not None for cell in row):
            return row[0].row
    return 1

def parse_date(date_str: str) -> str:
    """Parse date strings to ISO format, handling all known data entry variations:
      - Paren annotations:   'November 6, 2021 (No Show)'      → strip parens
      - Bare suffixes:       'November 6, 2021 Early Show'      → strip trailing words
      - CANCELLED:           'February 1, 2026 (Matinee) CANCELLED' → strip
      - Slash name suffix:   'August 1, 2021/Sorenson'          → take part before /
      - Period not comma:    'October 11. 2025'                 → replace . with ,
      - Missing space:       'December 28,2024'                 → add space
      - Extra comma:         'August, 17, 2024'                 → remove extra comma
      - Month typos:         'Februrary', 'Novmeber'            → correct spelling
    Returns None if the value is a summary label (e.g. 'July Financial Results').
    """
    import re
    clean = date_str.strip()

    # reject summary/label columns — not actual show dates
    if re.search(r'financial results|^june$|^total', clean, re.IGNORECASE):
        return None

    # take part before slash: "August 1, 2021/Sorenson" → "August 1, 2021"
    clean = clean.split("/")[0].strip()

    # strip paren annotations e.g. "(No Show)", "(Matinee)", "(Early)"
    clean = re.sub(r'\(.*?\)', '', clean).strip()

    # strip bare trailing words after the year: CANCELLED, Early, Late, Show, Matinee, Emerging Artist
    clean = re.sub(r'(\d{4})\s+.*$', r'\1', clean).strip()

    # replace period-as-comma: "October 11. 2025" → "October 11, 2025"
    clean = re.sub(r'(\d+)\.\s+(\d{4})', r'\1, \2', clean)

    # add missing space after comma: "December 28,2024" → "December 28, 2024"
    clean = re.sub(r'(\d),(\d)', r'\1, \2', clean)

    # remove extra comma after month name: "August, 17, 2024" → "August 17, 2024"
    clean = re.sub(r'([A-Za-z]+),\s+(\d)', r'\1 \2', clean)

    # fix month name typos
    clean = fix_month_typos(clean)

    for fmt in ("%B %d, %Y", "%B %d %Y", "%b %d, %Y", "%b %d %Y"):
        try:
            return datetime.strptime(clean.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None  # unparseable — exclude from DB

def load_file(path: Path, ws_name: str, ws, header_row: int,
              cursor, source_name: str):
    """Unpivot one sheet and insert rows into daily_values.
    Column A = metric label. Other columns with a parseable date = value columns."""
    headers = [cell.value for cell in ws[header_row]]

    # map col index → (person, iso_date) for date columns (skip col 0 = column A)
    value_cols = {}
    for idx, h in enumerate(headers):
        if idx == 0:
            continue
        if not isinstance(h, str):
            continue
        person, date_part = split_name_and_date(h)
        if date_part:
            iso = parse_date(date_part)
            if iso:
                value_cols[idx] = (person.strip(), iso)

    inserted = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue
        metric = str(row[0]).strip() if row[0] is not None else ""
        if not metric:
            continue

        for col_idx, (person, date_iso) in value_cols.items():
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if raw is None:
                continue
            try:
                value = float(raw)
            except (TypeError, ValueError):
                continue

            cursor.execute(
                "INSERT INTO daily_values (source_file, sheet, person, metric, date, value) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                [source_name, ws_name, person, metric, date_iso, value]
            )
            inserted += 1

    return inserted

def create_table(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS daily_values (
            id          INTEGER PRIMARY KEY,
            source_file TEXT,
            sheet       TEXT,
            person      TEXT,
            metric      TEXT,
            date        TEXT,
            value       REAL
        )
    """)

if __name__ == "__main__":
    if "--help" in sys.argv:
        print(HELP)
        sys.exit(0)

    args = sys.argv[1:]
    db_path = None
    i = 0
    while i < len(args):
        if args[i] == "--db" and i + 1 < len(args):
            db_path = Path(args[i + 1])
            i += 2
        else:
            i += 1

    mod_excel = Path("mod_excel")
    if not mod_excel.exists():
        print("mod_excel/ directory not found. Run anonymize_headers.py first.")
        sys.exit(1)

    file_paths = sorted(mod_excel.glob("*_anonymized.xlsx"))
    if not file_paths:
        print("No *_anonymized.xlsx files found in mod_excel/")
        sys.exit(1)

    if db_path is None:
        db_path = mod_excel / "financial_data.db"

    print(f"Found {len(file_paths)} anonymized file(s)")

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    create_table(cursor)

    total = 0
    for path in file_paths:
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"\nLoading: {path.name}")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = find_header_row(ws)
            n = load_file(path, sheet_name, ws, header_row,
                          cursor, path.stem)
            print(f"  tab: {sheet_name} — {n} rows inserted")
            total += n

    conn.commit()
    conn.close()

    print(f"\nDatabase saved: {db_path}")
    print(f"Total rows:     {total}")
